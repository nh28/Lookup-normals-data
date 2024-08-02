import logging
import openpyxl

class Table:
    def __init__(self, arkeon, user_elements, user_months):
        """
        Initialize a new instance of Table.

        Parameters:
        arkeon: An instance of the ARKEON object to be able to query data.
        user_elements: A list of elements that the user selected.
        user_months: A list of months that the user selected.
        """
        self.arkeon = arkeon
        self.user_elements = user_elements
        self.user_months = user_months
        self.normals_elements = self.arkeon.get_dataframe(['NORMAL_ID', 'E_NORMAL_ELEMENT_NAME'], 'NORMALS_1991.valid_normals_elements', [], False)
        self.extreme_el = self.arkeon.get_extreme_el()
        try:
            self.workbook = openpyxl.load_workbook('Input_Files/StationList.xlsx')
            self.worksheet = self.workbook["Sheet1"]
        except Exception as e:
            logging.error("Unable to find StationList.xlsx with Sheet1")
            print(e)

    def check_station_availability(self, station):
        """
        Finds the index in StationList.xlsx of the station that the user selected.

        Parameters:
        None

        Returns:
        row_idx: The index of the station list, and returns None if the station is not found.
        """
        for row_idx in range(2, self.worksheet.max_row):
            name_7181 = self.worksheet.cell(row = row_idx, column = 2).value
            name_91 = self.worksheet.cell(row = row_idx, column = 7).value
            if name_7181 == station or name_91 == station:
                return row_idx
        return None

    def format(self, extreme, df):
        """
        Formats the data based on the dataframe and if the element is an extreme element.

        Parameters:
        extreme: True if the given element is an extreme element.
        df: A dataframe that contains the data based on station, element, and month

        Returns:
        A formatted list with the information.
        """
        if df.empty:
            return [ "", "", ""]
        elif extreme:
            return [ df.iloc[0]["VALUE"],  df.iloc[0]["NORMAL_CODE"],  df.iloc[0]["FIRST_OCCURRENCE_DATE"]]
        else:
            return [ df.iloc[0]["VALUE"],  df.iloc[0]["NORMAL_CODE"],  ""]
    
    def get_all_data(self, station, master_df_71, master_df_81, master_df_91):
        """
        Loops through all the user elements and months to compile a dataframe containing all the data.

        Parameters:
        None

        Returns: 
        list: A list with all the data.
        """
        list = []
        list_81 = []
        list_91 = []
        
        row_idx = self.check_station_availability(station)
        if row_idx is not None:
            station_info = [cell.value for cell in self.worksheet[row_idx]]
            id_7181, name_7181, climate_id, exist_71, exist_81, id_91, name_91, prov = station_info[:8]

            for element in self.user_elements:
                for month in self.user_months:
                    normal_id = self.arkeon.get_normals_element_id(element, self.normals_elements)
                    extreme = element in self.extreme_el

                    data_71 = data_81 = data_91 = [""] * 3

                    if exist_71:
                        filtered_df_71 = master_df_71.loc[(master_df_71['STN_ID'] == int(id_7181)) & 
                                                        (master_df_71['NORMAL_ID'] == int(normal_id)) & 
                                                        (master_df_71['MONTH'] == int(month))]
                        data_71 = self.format(extreme, filtered_df_71)

                    if exist_81:
                        filtered_df_81 = master_df_81.loc[(master_df_81['STN_ID'] == int(id_7181)) & 
                                                        (master_df_81['NORMAL_ID'] == int(normal_id)) & 
                                                        (master_df_81['MONTH'] == int(month))]
                        data_81 = self.format(extreme, filtered_df_81)

                    if id_91:
                        filtered_df_91 = master_df_91.loc[(master_df_91['STN_ID'] == int(id_91)) & 
                                                        (master_df_91['NORMAL_ID'] == int(normal_id)) & 
                                                        (master_df_91['MONTH'] == int(month))]
                        data_91 = self.format(extreme, filtered_df_91)

                    row = [id_7181, name_7181, climate_id, prov, normal_id, element, month]        
                    row += data_71 + data_81 + data_91

                    if exist_71:
                        list.append(row)
                    elif exist_81:
                        list_81.append(row)
                    elif id_91:
                        row = [id_91, name_91, "", prov, normal_id, element, month] + data_71 + data_81 + data_91
                        list_91.append(row)

        list = list + list_81 + list_91 
        return list
        